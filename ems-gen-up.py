import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class EmployeeManagementUpdater:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Management System Updater")
        self.root.geometry("1000x700")
        
        self.file_path = ""
        self.workbook = None
        self.sheets = {}
        self.current_sheet = ""
        
        self.create_widgets()
        
    def create_widgets(self):
        # File selection frame
        file_frame = ttk.LabelFrame(self.root, text="Excel File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.file_entry = ttk.Entry(file_frame, width=50)
        self.file_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        browse_btn = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        browse_btn.pack(side=tk.LEFT, padx=5)
        
        load_btn = ttk.Button(file_frame, text="Load", command=self.load_workbook)
        load_btn.pack(side=tk.LEFT, padx=5)
        
        # Sheet selection
        sheet_frame = ttk.LabelFrame(self.root, text="Select Sheet", padding=10)
        sheet_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.sheet_combobox = ttk.Combobox(sheet_frame, state="readonly")
        self.sheet_combobox.pack(fill=tk.X, padx=5, pady=5)
        self.sheet_combobox.bind("<<ComboboxSelected>>", self.load_sheet_data)
        
        # Data display frame
        data_frame = ttk.LabelFrame(self.root, text="Sheet Data", padding=10)
        data_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Treeview for displaying data
        self.tree = ttk.Treeview(data_frame)
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(data_frame, orient="vertical", command=self.tree.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        x_scroll = ttk.Scrollbar(data_frame, orient="horizontal", command=self.tree.xview)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Edit frame
        edit_frame = ttk.LabelFrame(self.root, text="Edit Data", padding=10)
        edit_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.edit_btn = ttk.Button(edit_frame, text="Edit Selected", command=self.edit_selected)
        self.edit_btn.pack(side=tk.LEFT, padx=5)
        
        self.add_btn = ttk.Button(edit_frame, text="Add New", command=self.add_new)
        self.add_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_btn = ttk.Button(edit_frame, text="Delete Selected", command=self.delete_selected)
        self.delete_btn.pack(side=tk.LEFT, padx=5)
        
        self.save_btn = ttk.Button(edit_frame, text="Save Changes", command=self.save_changes)
        self.save_btn.pack(side=tk.RIGHT, padx=5)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.pack(fill=tk.X, padx=10, pady=5)
        
        # Disable buttons until file is loaded
        self.toggle_buttons(False)
    
    def toggle_buttons(self, enabled):
        state = tk.NORMAL if enabled else tk.DISABLED
        self.sheet_combobox.config(state=tk.NORMAL if enabled else tk.DISABLED)
        self.edit_btn.config(state=state)
        self.add_btn.config(state=state)
        self.delete_btn.config(state=state)
        self.save_btn.config(state=state)
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, file_path)
    
    def load_workbook(self):
        file_path = self.file_entry.get()
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file first.")
            return
        
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self.file_path = file_path
            self.sheets = {sheet.title: sheet for sheet in self.workbook.worksheets}
            
            # Update sheet combobox
            self.sheet_combobox["values"] = list(self.sheets.keys())
            if self.sheets:
                self.sheet_combobox.current(0)
                self.load_sheet_data()
            
            self.toggle_buttons(True)
            self.status_var.set(f"Loaded: {os.path.basename(file_path)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook:\n{str(e)}")
            self.status_var.set("Error loading file")
    
    def load_sheet_data(self, event=None):
        sheet_name = self.sheet_combobox.get()
        if not sheet_name or sheet_name not in self.sheets:
            return
        
        self.current_sheet = sheet_name
        sheet = self.sheets[sheet_name]
        
        # Clear the treeview
        self.tree.delete(*self.tree.get_children())
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Configure treeview columns
        self.tree["columns"] = headers
        self.tree.column("#0", width=0, stretch=tk.NO)  # Hide first empty column
        
        for header in headers:
            self.tree.column(header, width=100, anchor=tk.W)
            self.tree.heading(header, text=header, anchor=tk.W)
        
        # Add data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.tree.insert("", tk.END, values=row)
        
        self.status_var.set(f"Displaying: {sheet_name}")
    
    def edit_selected(self):
        selected = self.tree.focus()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row to edit.")
            return
        
        # Get selected item data
        item_data = self.tree.item(selected, "values")
        headers = self.tree["columns"]
        
        # Create edit window
        edit_win = tk.Toplevel(self.root)
        edit_win.title("Edit Record")
        edit_win.grab_set()
        
        # Create entry widgets for each column
        entries = {}
        for i, header in enumerate(headers):
            ttk.Label(edit_win, text=header).grid(row=i, column=0, padx=5, pady=2, sticky=tk.E)
            entry = ttk.Entry(edit_win)
            entry.grid(row=i, column=1, padx=5, pady=2, sticky=tk.W+tk.E)
            
            # Pre-fill with current value
            if i < len(item_data):
                entry.insert(0, str(item_data[i]) if item_data[i] is not None else "")
            
            entries[header] = entry
        
        # Save button
        save_btn = ttk.Button(
            edit_win, 
            text="Save Changes",
            command=lambda: self.save_edit(selected, entries, edit_win)
        )
        save_btn.grid(row=len(headers), column=0, columnspan=2, pady=10)
    
    def save_edit(self, item_id, entries, window):
        # Get new values from entries
        new_values = []
        for header in self.tree["columns"]:
            new_values.append(entries[header].get())
        
        # Update treeview
        self.tree.item(item_id, values=new_values)
        
        # Close edit window
        window.destroy()
        self.status_var.set("Changes saved to memory (not file). Click 'Save Changes' to update file.")
    
    def add_new(self):
        headers = self.tree["columns"]
        
        # Create add window
        add_win = tk.Toplevel(self.root)
        add_win.title("Add New Record")
        add_win.grab_set()
        
        # Create entry widgets for each column
        entries = {}
        for i, header in enumerate(headers):
            ttk.Label(add_win, text=header).grid(row=i, column=0, padx=5, pady=2, sticky=tk.E)
            entry = ttk.Entry(add_win)
            entry.grid(row=i, column=1, padx=5, pady=2, sticky=tk.W+tk.E)
            entries[header] = entry
        
        # Add button
        add_btn = ttk.Button(
            add_win, 
            text="Add Record",
            command=lambda: self.save_new(entries, add_win)
        )
        add_btn.grid(row=len(headers), column=0, columnspan=2, pady=10)
    
    def save_new(self, entries, window):
        # Get new values from entries
        new_values = []
        for header in self.tree["columns"]:
            new_values.append(entries[header].get())
        
        # Add to treeview
        self.tree.insert("", tk.END, values=new_values)
        
        # Close add window
        window.destroy()
        self.status_var.set("New record added to memory (not file). Click 'Save Changes' to update file.")
    
    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select one or more rows to delete.")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete the selected records?"):
            for item in selected:
                self.tree.delete(item)
            self.status_var.set("Records deleted from memory (not file). Click 'Save Changes' to update file.")
    
    def save_changes(self):
        if not self.workbook or not self.current_sheet:
            messagebox.showerror("Error", "No workbook or sheet loaded.")
            return
        
        try:
            sheet = self.workbook[self.current_sheet]
            
            # Clear existing data (keep headers)
            sheet.delete_rows(2, sheet.max_row)
            
            # Get headers
            headers = self.tree["columns"]
            
            # Add data from treeview
            for item in self.tree.get_children():
                values = self.tree.item(item, "values")
                sheet.append(values)
            
            # Save the workbook
            self.workbook.save(self.file_path)
            self.status_var.set(f"Changes saved to {os.path.basename(self.file_path)}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save changes:\n{str(e)}")
            self.status_var.set("Error saving changes")

if __name__ == "__main__":
    root = tk.Tk()
    app = EmployeeManagementUpdater(root)
    root.mainloop()